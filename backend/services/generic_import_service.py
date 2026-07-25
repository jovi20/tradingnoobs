"""Persistent, owner-bound generic import upload and preview support."""
from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
import hashlib
import json
import os
from pathlib import Path
import tempfile
from typing import Any, Iterable

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy import update
from sqlalchemy.orm import Session

from app_config.release_contract import (
    JOURNAL_BETA_CONTRACT,
    ReleaseContractViolation,
    require_release_currency,
)
from models import (
    AssetMaster,
    ImportAdapterKind,
    ImportRow,
    ImportSession,
    ImportSessionStatus,
    TradeInstrument,
    TradingAccount,
)
from services.financial_command_service import (
    permanently_forbid_account_hard_delete,
)
from services.idempotency_service import (
    begin_idempotent_request,
    complete_idempotent_request,
)
from services.instrument_identity_service import (
    InstrumentIdentity,
    canonical_asset_code,
)
from services.legacy_truth_sync_service import validate_legacy_instrument_identity
from services.timezone_service import (
    LocalDateTimeError,
    normalize_user_datetime_to_utc,
)


UPLOAD_SCOPE = "GENERIC_IMPORT_UPLOAD_V1"
RESPONSE_SCHEMA_VERSION = 1
IMPORT_FILE_PREFIX = "tradingnoobs-import-"
MAX_FILE_BYTES = JOURNAL_BETA_CONTRACT.imports.common_limits.max_file_bytes
MAX_ROWS = JOURNAL_BETA_CONTRACT.imports.common_limits.max_rows_or_executions
PREVIEW_TTL_SECONDS = (
    JOURNAL_BETA_CONTRACT.imports.common_limits.preview_ttl_seconds
)
ROW_RETENTION_DAYS = (
    JOURNAL_BETA_CONTRACT.imports.common_limits.terminal_normalized_row_retention_days
)
TERMINAL_STATUSES = frozenset(
    JOURNAL_BETA_CONTRACT.source_states.terminal_import_session_states
)
NONTERMINAL_STATUSES = frozenset(
    set(JOURNAL_BETA_CONTRACT.source_states.import_session_states)
    - TERMINAL_STATUSES
)

REQUIRED_COLUMNS = (
    "asset_type",
    "market",
    "exchange_code",
    "symbol",
    "instrument_type",
    "direction",
    "action",
    "timestamp",
    "price",
    "quantity",
    "currency",
)
OPTIONAL_COLUMNS = (
    "commission",
    "fee_currency",
    "reason",
    "note",
)
UNTRUSTED_SOURCE_COLUMNS = frozenset(
    {
        "external_trade_id",
        "external_execution_id",
        "execution_id",
        "source_event_id",
        "source_id",
        "trade_id",
        "order_id",
    }
)
COLUMN_ALIASES = {
    "asset": "asset_type",
    "core_type": "asset_type",
    "asset class": "asset_type",
    "venue": "market",
    "exchange": "exchange_code",
    "ticker": "symbol",
    "code": "symbol",
    "instrument": "instrument_type",
    "side": "direction",
    "operation": "action",
    "event_type": "action",
    "event type": "action",
    "date": "timestamp",
    "time": "timestamp",
    "datetime": "timestamp",
    "occurred_at": "timestamp",
    "occurred at": "timestamp",
    "trade_time": "timestamp",
    "trade time": "timestamp",
    "trade_price": "price",
    "trade price": "price",
    "cost": "price",
    "qty": "quantity",
    "amount": "quantity",
    "quote_currency": "currency",
    "quote currency": "currency",
    "fee": "commission",
    "comm": "commission",
}
DIRECTION_ALIASES = {
    "LONG": "LONG",
    "L": "LONG",
    "BUY": "LONG",
    "SHORT": "SHORT",
    "S": "SHORT",
    "SELL": "SHORT",
}
ACTION_ALIASES = {
    "OPEN": "OPEN",
    "ENTRY": "OPEN",
    "ADD": "ADD",
    "REDUCE": "REDUCE",
    "CLOSE": "CLOSE",
}


class GenericImportError(ValueError):
    def __init__(self, code: str, message: str, *, http_status: int = 422):
        self.code = code
        self.http_status = http_status
        super().__init__(message)


@dataclass(frozen=True)
class StagedImportFile:
    path: Path
    file_hash: str
    size_bytes: int
    file_format: str
    original_filename: str
    media_type: str | None


@dataclass(frozen=True)
class UploadCommandResult:
    body: dict[str, Any]
    http_status: int
    replayed: bool


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _as_utc(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def import_temp_root() -> Path:
    configured = os.getenv("TRADINGNOOBS_IMPORT_TMP_DIR")
    root = (
        Path(configured)
        if configured
        else Path(tempfile.gettempdir()) / "tradingnoobs-imports"
    )
    root.mkdir(mode=0o700, parents=True, exist_ok=True)
    os.chmod(root, 0o700)
    return root


def _file_format(filename: str) -> tuple[str, str]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return "CSV_UTF8", ".csv"
    if suffix == ".xlsx":
        return "XLSX", ".xlsx"
    raise GenericImportError(
        "UNSUPPORTED_IMPORT_FORMAT",
        "Only UTF-8 CSV and XLSX files are supported",
    )


def _safe_filename(value: str | None) -> str:
    filename = Path((value or "import").replace("\x00", "")).name.strip()
    return (filename or "import")[:255]


async def stage_import_upload(
    upload: UploadFile,
    *,
    temp_root: Path | None = None,
) -> StagedImportFile:
    original_filename = _safe_filename(upload.filename)
    file_format, suffix = _file_format(original_filename)
    root = temp_root or import_temp_root()
    root.mkdir(mode=0o700, parents=True, exist_ok=True)
    os.chmod(root, 0o700)
    descriptor, raw_path = tempfile.mkstemp(
        prefix=IMPORT_FILE_PREFIX,
        suffix=suffix,
        dir=root,
    )
    path = Path(raw_path)
    os.chmod(path, 0o600)
    digest = hashlib.sha256()
    size = 0
    try:
        with os.fdopen(descriptor, "wb") as staged:
            while True:
                chunk = await upload.read(1024 * 1024)
                if not chunk:
                    break
                size += len(chunk)
                if size > MAX_FILE_BYTES:
                    raise GenericImportError(
                        "IMPORT_FILE_TOO_LARGE",
                        f"Import file must not exceed {MAX_FILE_BYTES} bytes",
                        http_status=413,
                    )
                digest.update(chunk)
                staged.write(chunk)
        return StagedImportFile(
            path=path,
            file_hash="sha256:" + digest.hexdigest(),
            size_bytes=size,
            file_format=file_format,
            original_filename=original_filename,
            media_type=upload.content_type,
        )
    except BaseException:
        path.unlink(missing_ok=True)
        raise


def remove_staged_import_file(staged: StagedImportFile | None) -> None:
    if staged is not None:
        staged.path.unlink(missing_ok=True)


def scavenge_orphan_import_files(
    *,
    now: datetime | None = None,
    older_than_seconds: int = 3600,
    temp_root: Path | None = None,
) -> int:
    root = temp_root or import_temp_root()
    threshold = _as_utc(now or utc_now()).timestamp() - older_than_seconds
    removed = 0
    for candidate in root.glob(f"{IMPORT_FILE_PREFIX}*"):
        try:
            if candidate.is_file() and candidate.stat().st_mtime <= threshold:
                candidate.unlink()
                removed += 1
        except FileNotFoundError:
            continue
    return removed


def _issue(code: str, message: str, *, field: str | None = None) -> dict[str, Any]:
    return {"code": code, "field": field, "message": message}


def _json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _normalized_header(value: Any) -> str:
    raw = str(value or "").strip().lower().replace("_", " ")
    return " ".join(raw.split())


def _canonical_headers(headers: Iterable[Any]) -> list[str]:
    canonical: list[str] = []
    for header in headers:
        normalized = _normalized_header(header)
        if not normalized:
            raise GenericImportError(
                "INVALID_IMPORT_HEADER",
                "Import headers must not be blank",
            )
        canonical_name = COLUMN_ALIASES.get(normalized, normalized.replace(" ", "_"))
        canonical.append(canonical_name)
    duplicates = sorted(
        {header for header in canonical if canonical.count(header) > 1}
    )
    if duplicates:
        raise GenericImportError(
            "DUPLICATE_IMPORT_HEADER",
            "Multiple columns resolve to the same canonical field: "
            + ", ".join(duplicates),
        )
    missing = sorted(set(REQUIRED_COLUMNS) - set(canonical))
    if missing:
        raise GenericImportError(
            "MISSING_IMPORT_COLUMNS",
            "Missing required columns: " + ", ".join(missing),
        )
    return canonical


def _read_csv_rows(path: Path) -> list[tuple[int, dict[str, Any]]]:
    try:
        handle = path.open("r", encoding="utf-8-sig", newline="")
    except UnicodeDecodeError as exc:
        raise GenericImportError(
            "INVALID_CSV_ENCODING",
            "CSV files must use UTF-8 encoding",
        ) from exc
    try:
        reader = csv.reader(handle)
        try:
            headers = next(reader)
        except StopIteration as exc:
            raise GenericImportError(
                "IMPORT_FILE_EMPTY",
                "Import file must include a canonical header row",
            ) from exc
        except UnicodeDecodeError as exc:
            raise GenericImportError(
                "INVALID_CSV_ENCODING",
                "CSV files must use UTF-8 encoding",
            ) from exc
        except csv.Error as exc:
            raise GenericImportError(
                "INVALID_CSV_FILE",
                "CSV file could not be parsed",
            ) from exc
        canonical = _canonical_headers(headers)
        rows: list[tuple[int, dict[str, Any]]] = []
        try:
            for row_number, values in enumerate(reader, start=2):
                if len(rows) >= MAX_ROWS:
                    raise GenericImportError(
                        "IMPORT_ROW_LIMIT_EXCEEDED",
                        f"Import file must not contain more than {MAX_ROWS} rows",
                    )
                if len(values) > len(canonical):
                    raise GenericImportError(
                        "IMPORT_ROW_COLUMN_MISMATCH",
                        f"Row {row_number} contains more values than the header",
                    )
                padded = values + [None] * (len(canonical) - len(values))
                rows.append((row_number, dict(zip(canonical, padded))))
        except UnicodeDecodeError as exc:
            raise GenericImportError(
                "INVALID_CSV_ENCODING",
                "CSV files must use UTF-8 encoding",
            ) from exc
        except csv.Error as exc:
            raise GenericImportError(
                "INVALID_CSV_FILE",
                "CSV file could not be parsed",
            ) from exc
        return rows
    finally:
        handle.close()


def _read_xlsx_rows(path: Path) -> list[tuple[int, dict[str, Any]]]:
    workbook = None
    try:
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=True,
        )
        worksheet = workbook.worksheets[0]
        iterator = worksheet.iter_rows(values_only=True)
        try:
            headers = next(iterator)
        except StopIteration as exc:
            raise GenericImportError(
                "IMPORT_FILE_EMPTY",
                "Import file must include a canonical header row",
            ) from exc
        canonical = _canonical_headers(headers)
        rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, values_tuple in enumerate(iterator, start=2):
            if len(rows) >= MAX_ROWS:
                raise GenericImportError(
                    "IMPORT_ROW_LIMIT_EXCEEDED",
                    f"Import file must not contain more than {MAX_ROWS} rows",
                )
            values = list(values_tuple)
            if len(values) > len(canonical):
                extra = values[len(canonical):]
                if any(value is not None for value in extra):
                    raise GenericImportError(
                        "IMPORT_ROW_COLUMN_MISMATCH",
                        f"Row {row_number} contains more values than the header",
                    )
                values = values[:len(canonical)]
            values += [None] * (len(canonical) - len(values))
            rows.append((row_number, dict(zip(canonical, values))))
        return rows
    except GenericImportError:
        raise
    except Exception as exc:
        raise GenericImportError(
            "INVALID_XLSX_FILE",
            "XLSX file could not be parsed",
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()


def read_staged_rows(staged: StagedImportFile) -> list[tuple[int, dict[str, Any]]]:
    if staged.file_format == "CSV_UTF8":
        return _read_csv_rows(staged.path)
    if staged.file_format == "XLSX":
        return _read_xlsx_rows(staged.path)
    raise GenericImportError(
        "UNSUPPORTED_IMPORT_FORMAT",
        "Only UTF-8 CSV and XLSX files are supported",
    )


def _positive_decimal(
    value: Any,
    *,
    field: str,
    allow_zero: bool = False,
) -> tuple[str | None, dict[str, Any] | None]:
    try:
        number = Decimal(str(value).strip())
    except (InvalidOperation, AttributeError, ValueError):
        return None, _issue(
            "INVALID_IMPORT_NUMBER",
            f"{field} must be a decimal number",
            field=field,
        )
    if not number.is_finite() or number < 0 or (number == 0 and not allow_zero):
        return None, _issue(
            "INVALID_IMPORT_NUMBER",
            f"{field} must be {'non-negative' if allow_zero else 'greater than zero'}",
            field=field,
        )
    return format(number, "f"), None


def _timestamp(
    value: Any,
    *,
    timezone_name: str,
) -> tuple[str | None, dict[str, Any] | None]:
    try:
        if isinstance(value, datetime):
            parsed = value
        elif isinstance(value, date):
            parsed = datetime.combine(value, datetime.min.time())
        else:
            token = str(value or "").strip()
            if token.endswith("Z"):
                token = token[:-1] + "+00:00"
            parsed = datetime.fromisoformat(token)
        normalized = normalize_user_datetime_to_utc(
            parsed,
            timezone_name=timezone_name,
        )
    except LocalDateTimeError as exc:
        return None, _issue(exc.code, str(exc), field="timestamp")
    except (TypeError, ValueError):
        return None, _issue(
            "INVALID_IMPORT_TIMESTAMP",
            "timestamp must be an ISO-8601 date-time",
            field="timestamp",
        )
    return normalized.isoformat().replace("+00:00", "Z"), None


def _instrument_exists(db: Session, identity: InstrumentIdentity) -> bool:
    asset = db.query(AssetMaster).filter(
        AssetMaster.canonical_code == canonical_asset_code(identity),
    ).first()
    if asset is None:
        return False
    return db.query(TradeInstrument.id).filter(
        TradeInstrument.asset_id == asset.id,
        TradeInstrument.contract_symbol == identity.normalized_symbol,
    ).first() is not None


def normalize_import_row(
    db: Session,
    *,
    raw_values: dict[str, Any],
    account: TradingAccount,
    timezone_name: str,
) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    raw = {key: _json_value(value) for key, value in raw_values.items()}
    normalized: dict[str, Any] = {}
    errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []

    direction_token = str(raw_values.get("direction") or "").strip().upper()
    direction = DIRECTION_ALIASES.get(direction_token)
    if direction is None:
        errors.append(
            _issue(
                "UNSUPPORTED_IMPORT_DIRECTION",
                "direction must resolve to LONG or SHORT",
                field="direction",
            )
        )
    else:
        normalized["direction"] = direction

    action_token = str(raw_values.get("action") or "").strip().upper()
    action = ACTION_ALIASES.get(action_token)
    if action is None:
        errors.append(
            _issue(
                "UNSUPPORTED_IMPORT_ACTION",
                "action must resolve to OPEN, ADD, REDUCE, or CLOSE",
                field="action",
            )
        )
    else:
        normalized["action"] = action

    occurred_at, timestamp_error = _timestamp(
        raw_values.get("timestamp"),
        timezone_name=timezone_name,
    )
    if timestamp_error:
        errors.append(timestamp_error)
    else:
        normalized["occurred_at"] = occurred_at

    price, price_error = _positive_decimal(raw_values.get("price"), field="price")
    quantity, quantity_error = _positive_decimal(
        raw_values.get("quantity"),
        field="quantity",
    )
    if price_error:
        errors.append(price_error)
    else:
        normalized["price"] = price
    if quantity_error:
        errors.append(quantity_error)
    else:
        normalized["quantity"] = quantity

    commission_value = raw_values.get("commission")
    if commission_value not in (None, ""):
        commission, commission_error = _positive_decimal(
            commission_value,
            field="commission",
            allow_zero=True,
        )
        if commission_error:
            errors.append(commission_error)
        else:
            normalized["commission"] = commission
    else:
        normalized["commission"] = "0"

    try:
        legacy_identity = validate_legacy_instrument_identity(
            position_asset_type=raw_values.get("asset_type"),
            account_currency=account.currency,
            symbol=raw_values.get("symbol"),
            exchange_code=raw_values.get("exchange_code"),
            metadata_core_type=raw_values.get("asset_type"),
            metadata_market=raw_values.get("market"),
            metadata_currency=raw_values.get("currency"),
            metadata_instrument=raw_values.get("instrument_type"),
        )
        identity = InstrumentIdentity(
            asset_type=legacy_identity.asset_type,
            market=legacy_identity.market,
            exchange_code=legacy_identity.exchange_code,
            normalized_symbol=legacy_identity.normalized_symbol,
            instrument_type=legacy_identity.instrument_type,
            quote_currency=legacy_identity.quote_currency,
        )
        normalized.update(
            {
                "asset_type": identity.asset_type,
                "market": identity.market,
                "exchange_code": identity.exchange_code,
                "symbol": identity.normalized_symbol,
                "instrument_type": identity.instrument_type,
                "currency": identity.quote_currency,
                "instrument_resolution": (
                    "EXISTING"
                    if _instrument_exists(db, identity)
                    else "CREATE_ON_CONFIRM"
                ),
            }
        )
    except ReleaseContractViolation as exc:
        errors.append(
            _issue(
                exc.code,
                "Instrument identity is outside the release allowlist",
                field=exc.field,
            )
        )

    fee_currency_value = raw_values.get("fee_currency")
    if fee_currency_value not in (None, ""):
        fee_currency = str(fee_currency_value).strip().upper()
        if fee_currency != account.currency:
            errors.append(
                _issue(
                    "INSTRUMENT_IDENTITY_MISMATCH",
                    "fee_currency must equal account currency",
                    field="fee_currency",
                )
            )
        else:
            normalized["fee_currency"] = fee_currency
    else:
        normalized["fee_currency"] = account.currency

    for optional in ("reason", "note"):
        value = raw_values.get(optional)
        if value not in (None, ""):
            normalized[optional] = str(value)

    claimed_source_fields = sorted(
        column
        for column in raw_values
        if column in UNTRUSTED_SOURCE_COLUMNS and raw_values[column] not in (None, "")
    )
    if claimed_source_fields:
        warnings.append(
            _issue(
                "UNTRUSTED_SOURCE_ID_IGNORED",
                "Generic files cannot declare trusted source identity",
                field=",".join(claimed_source_fields),
            )
        )
    return normalized, errors, warnings


def _row_fingerprint(normalized: dict[str, Any]) -> str:
    serialized = json.dumps(
        normalized,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=True,
    )
    return hashlib.sha256(serialized.encode("ascii")).hexdigest()


def persist_preview_rows(
    db: Session,
    *,
    session: ImportSession,
    account: TradingAccount,
    timezone_name: str,
    rows: list[tuple[int, dict[str, Any]]],
) -> bool:
    seen: dict[str, int] = {}
    has_dst_error = False
    valid_rows = 0
    error_rows = 0
    warning_rows = 0
    for row_number, raw_values in rows:
        normalized, errors, warnings = normalize_import_row(
            db,
            raw_values=raw_values,
            account=account,
            timezone_name=timezone_name,
        )
        fingerprint = _row_fingerprint(normalized)
        if fingerprint in seen:
            warnings.append(
                _issue(
                    "DUPLICATE_ROW",
                    f"Normalized row duplicates row {seen[fingerprint]}; it is not removed",
                )
            )
        else:
            seen[fingerprint] = row_number
        is_valid = not errors
        valid_rows += int(is_valid)
        error_rows += int(not is_valid)
        warning_rows += int(bool(warnings))
        has_dst_error = has_dst_error or any(
            issue["code"] in {"AMBIGUOUS_LOCAL_TIME", "NONEXISTENT_LOCAL_TIME"}
            for issue in errors
        )
        db.add(
            ImportRow(
                session=session,
                user_id=session.user_id,
                account_id=session.account_id,
                adapter_kind=session.adapter_kind,
                file_hash=session.file_hash,
                row_number=row_number,
                raw_values_json={
                    key: _json_value(value)
                    for key, value in raw_values.items()
                },
                normalized_values_json=normalized,
                validation_errors_json=errors,
                warnings_json=warnings,
                is_valid=is_valid,
            )
        )
    session.total_rows = len(rows)
    session.valid_rows = valid_rows
    session.error_rows = error_rows
    session.warning_rows = warning_rows
    db.flush()
    return has_dst_error


def transition_import_session(
    db: Session,
    *,
    session: ImportSession,
    from_status: str,
    to_status: str,
    now: datetime | None = None,
) -> None:
    allowed = {
        item.from_state: set(item.to_states)
        for item in JOURNAL_BETA_CONTRACT.source_states.import_session_transitions
    }
    if to_status not in allowed.get(from_status, set()):
        raise ValueError(f"Illegal ImportSession transition {from_status} -> {to_status}")
    now = _as_utc(now or utc_now())
    values: dict[str, Any] = {
        "status": to_status,
        "status_changed_at": now,
    }
    if to_status in TERMINAL_STATUSES:
        values["terminal_at"] = now
    result = db.execute(
        update(ImportSession)
        .where(
            ImportSession.id == session.id,
            ImportSession.status == from_status,
        )
        .values(**values)
    )
    if result.rowcount != 1:
        raise GenericImportError(
            "IMPORT_SESSION_STATE_CONFLICT",
            "Import session state changed concurrently",
            http_status=409,
        )
    for key, value in values.items():
        setattr(session, key, value)
    db.flush()


def _serialize_row(row: ImportRow) -> dict[str, Any]:
    return {
        "public_id": row.public_id,
        "row_number": row.row_number,
        "raw_values": row.raw_values_json or {},
        "normalized_values": row.normalized_values_json or {},
        "is_valid": bool(row.is_valid),
        "errors": row.validation_errors_json or [],
        "warnings": row.warnings_json or [],
    }


def serialize_import_session(
    db: Session,
    *,
    session: ImportSession,
    include_rows: bool,
) -> dict[str, Any]:
    account_public_id = db.query(TradingAccount.public_id).filter(
        TradingAccount.id == session.account_id,
        TradingAccount.user_id == session.user_id,
    ).scalar()
    rows = []
    if include_rows:
        rows = (
            db.query(ImportRow)
            .filter(
                ImportRow.session_id == session.id,
                ImportRow.user_id == session.user_id,
            )
            .order_by(ImportRow.row_number.asc())
            .all()
        )
    error = None
    if session.error_code:
        error = _issue(
            session.error_code,
            session.error_message or "Import upload failed",
        )
    return {
        "schema_version": RESPONSE_SCHEMA_VERSION,
        "session_public_id": session.public_id,
        "account_public_id": account_public_id,
        "adapter_kind": session.adapter_kind,
        "file_format": session.file_format,
        "status": session.status,
        "expires_at": _as_utc(session.expires_at).isoformat().replace(
            "+00:00",
            "Z",
        ),
        "total_rows": session.total_rows,
        "valid_rows": session.valid_rows,
        "error_rows": session.error_rows,
        "warning_rows": session.warning_rows,
        "error": error,
        "rows": [_serialize_row(row) for row in rows],
        "confirm_available": False,
    }


def _idempotency_key_hash(raw_key: str | None) -> str:
    normalized = (raw_key or "").strip()
    if not normalized:
        raise GenericImportError(
            "IDEMPOTENCY_KEY_REQUIRED",
            "Idempotency-Key is required for import upload",
        )
    if len(normalized) > 255:
        raise GenericImportError(
            "IDEMPOTENCY_KEY_INVALID",
            "Idempotency-Key must be at most 255 characters",
        )
    return "sha256:" + hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def upload_preview(
    db: Session,
    *,
    user_id: int,
    timezone_name: str,
    account: TradingAccount,
    staged: StagedImportFile,
    idempotency_key: str | None,
    now: datetime | None = None,
) -> UploadCommandResult:
    now = _as_utc(now or utc_now())
    key_hash = _idempotency_key_hash(idempotency_key)
    request_payload = {
        "account_public_id": account.public_id,
        "adapter_kind": ImportAdapterKind.GENERIC_BOOTSTRAP.value,
        "file_hash": staged.file_hash,
    }
    try:
        command = begin_idempotent_request(
            db,
            scope=UPLOAD_SCOPE,
            key=key_hash,
            request_payload=request_payload,
            user_id=user_id,
            ttl_seconds=None,
            now=now,
        )
    except ValueError as exc:
        raise GenericImportError(
            "IDEMPOTENCY_KEY_REUSED",
            str(exc),
            http_status=409,
        ) from exc
    if not command.created:
        if (
            command.record.status == "COMPLETED"
            and isinstance(command.record.response_json, dict)
        ):
            envelope = command.record.response_json
            return UploadCommandResult(
                body=envelope["body"],
                http_status=int(envelope["http_status"]),
                replayed=True,
            )
        raise GenericImportError(
            "IDEMPOTENCY_REQUEST_IN_PROGRESS",
            "Import upload with this Idempotency-Key is already in progress",
            http_status=409,
        )

    if not account.is_active:
        raise GenericImportError(
            "ACCOUNT_ARCHIVED",
            "Archived accounts reject new import sessions",
            http_status=409,
        )
    try:
        require_release_currency(account.currency, field="account.currency")
    except ReleaseContractViolation as exc:
        raise GenericImportError(
            exc.code,
            "Account currency is outside the release contract",
            http_status=422,
        ) from exc

    session = ImportSession(
        user_id=user_id,
        account_id=account.id,
        upload_idempotency_id=command.record.id,
        adapter_kind=ImportAdapterKind.GENERIC_BOOTSTRAP.value,
        file_format=staged.file_format,
        file_hash=staged.file_hash,
        file_size_bytes=staged.size_bytes,
        original_filename=staged.original_filename,
        media_type=staged.media_type,
        status=ImportSessionStatus.UPLOADING.value,
        expires_at=now + timedelta(seconds=PREVIEW_TTL_SECONDS),
        status_changed_at=now,
        response_schema_version=RESPONSE_SCHEMA_VERSION,
    )
    db.add(session)
    permanently_forbid_account_hard_delete(account)
    db.flush()

    http_status = 201
    try:
        rows = read_staged_rows(staged)
        has_dst_error = persist_preview_rows(
            db,
            session=session,
            account=account,
            timezone_name=timezone_name,
            rows=rows,
        )
        transition_import_session(
            db,
            session=session,
            from_status=ImportSessionStatus.UPLOADING.value,
            to_status=ImportSessionStatus.PREVIEW_READY.value,
            now=now,
        )
        if has_dst_error:
            http_status = 422
    except GenericImportError as exc:
        session.error_code = exc.code
        session.error_message = str(exc)
        transition_import_session(
            db,
            session=session,
            from_status=ImportSessionStatus.UPLOADING.value,
            to_status=ImportSessionStatus.FAILED.value,
            now=now,
        )
        http_status = exc.http_status

    body = serialize_import_session(
        db,
        session=session,
        include_rows=session.status == ImportSessionStatus.PREVIEW_READY.value,
    )
    complete_idempotent_request(
        db,
        record=command.record,
        response_json={"http_status": http_status, "body": body},
        source_fact_public_id=session.public_id,
        now=now,
    )
    db.commit()
    return UploadCommandResult(body=body, http_status=http_status, replayed=False)


def get_owned_import_session(
    db: Session,
    *,
    user_id: int,
    session_public_id: str,
) -> ImportSession | None:
    return db.query(ImportSession).filter(
        ImportSession.public_id == session_public_id,
        ImportSession.user_id == user_id,
    ).first()


def expire_session_if_due(
    db: Session,
    *,
    session: ImportSession,
    now: datetime | None = None,
) -> bool:
    now = _as_utc(now or utc_now())
    if (
        session.status in {
            ImportSessionStatus.UPLOADING.value,
            ImportSessionStatus.PREVIEW_READY.value,
        }
        and _as_utc(session.expires_at) <= now
    ):
        transition_import_session(
            db,
            session=session,
            from_status=session.status,
            to_status=ImportSessionStatus.EXPIRED.value,
            now=now,
        )
        db.commit()
        return True
    return session.status == ImportSessionStatus.EXPIRED.value


def expire_due_import_sessions(
    db: Session,
    *,
    now: datetime | None = None,
    batch_size: int = 100,
) -> int:
    now = _as_utc(now or utc_now())
    candidates = (
        db.query(ImportSession)
        .filter(
            ImportSession.status.in_(
                (
                    ImportSessionStatus.UPLOADING.value,
                    ImportSessionStatus.PREVIEW_READY.value,
                )
            ),
            ImportSession.expires_at <= now,
        )
        .order_by(ImportSession.expires_at.asc(), ImportSession.id.asc())
        .limit(batch_size)
        .all()
    )
    expired = 0
    for session in candidates:
        try:
            transition_import_session(
                db,
                session=session,
                from_status=session.status,
                to_status=ImportSessionStatus.EXPIRED.value,
                now=now,
            )
            expired += 1
        except GenericImportError:
            db.expire(session)
    db.flush()
    return expired


def cleanup_terminal_import_rows(
    db: Session,
    *,
    now: datetime | None = None,
    batch_size: int = 1000,
) -> int:
    now = _as_utc(now or utc_now())
    cutoff = now - timedelta(days=ROW_RETENTION_DAYS)
    rows = (
        db.query(ImportRow.id, ImportRow.session_id)
        .join(ImportSession, ImportSession.id == ImportRow.session_id)
        .filter(
            ImportSession.status.in_(tuple(TERMINAL_STATUSES)),
            ImportSession.terminal_at.is_not(None),
            ImportSession.terminal_at <= cutoff,
        )
        .order_by(ImportRow.id.asc())
        .limit(batch_size)
        .all()
    )
    if not rows:
        return 0
    row_ids = [row_id for row_id, _session_id in rows]
    session_ids = sorted({session_id for _row_id, session_id in rows})
    deleted = (
        db.query(ImportRow)
        .filter(ImportRow.id.in_(row_ids))
        .delete(synchronize_session=False)
    )
    for session_id in session_ids:
        has_rows = db.query(ImportRow.id).filter(
            ImportRow.session_id == session_id,
        ).first()
        if has_rows is None:
            db.query(ImportSession).filter(
                ImportSession.id == session_id,
                ImportSession.rows_cleaned_at.is_(None),
            ).update(
                {"rows_cleaned_at": now},
                synchronize_session=False,
            )
    db.flush()
    return deleted
